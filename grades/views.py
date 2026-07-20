import io
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.db.models import Q
from core.decorators import tenant_required, admin_or_teacher_required
from core.activity import log_activity
from accounts.utils import is_admin
from scheduling.models import AcademicYear, Enrolment, Section
from students.models import Student
from .forms import BulkGradeUploadForm, EvaluationForm, GradeVisibilityForm
from .models import Evaluation, GradeComment, GradeEntry, GradeWindow, ReportCard
from .utils import compute_student_average, grade_window_is_open


def get_teacher_staff(user):
    """Get the staff profile associated with a user"""
    try:
        return user.staff_profile
    except Exception:
        return None


def get_hod_departments(staff):
    """Get departments this staff member heads (if flagged as HOD)"""
    if staff and staff.is_head_of_department:
        return [d for d in (staff.department, staff.department_2) if d]
    return []


def section_in_departments(section, departments):
    """Check if a section belongs to any of the given departments"""
    if not departments:
        return False
    dept = section.course.department
    return bool(dept) and dept in departments


# -- Grades Home ---------------------------------------------------------------


@admin_or_teacher_required
@tenant_required
def grades_home(request):
    school = request.school
    admin = is_admin(request.user, school)
    staff = get_teacher_staff(request.user)

    if admin:
        sections = Section.objects.filter(school=school).select_related("course", "form", "academic_year", "teacher")
    elif staff:
        hod_departments = get_hod_departments(staff)
        if hod_departments:
            sections = (
                Section.objects.filter(school=school)
                .filter(Q(teacher=staff) | Q(course__department__in=hod_departments))
                .select_related("course", "form", "academic_year", "teacher")
            )
        else:
            sections = Section.objects.filter(school=school, teacher=staff).select_related("course", "form", "academic_year", "teacher")
    else:
        messages.error(request, "Access denied.")
        return redirect("portals:dashboard")

    years = AcademicYear.objects.filter(school=school).order_by("-name")
    year_pk = request.GET.get("year")
    term = request.GET.get("term")

    if year_pk:
        sections = sections.filter(academic_year_id=year_pk)
    if term:
        sections = sections.filter(term_number=term)

    window_statuses = {}
    for section in sections:
        window_statuses[section.pk] = grade_window_is_open(school, section.academic_year, section.term_number, section.form)

    return render(
        request,
        "grades/home.html",
        {
            "sections": sections,
            "years": years,
            "selected_year": year_pk,
            "selected_term": term,
            "is_admin": admin,
            "window_statuses": window_statuses,
            "current_staff": staff,
        },
    )


# -- Section Grade Table -------------------------------------------------------
@login_required
@tenant_required
def section_grade_table(request, section_pk):
    section = get_object_or_404(Section, pk=section_pk, school=request.school)
    school = request.school
    admin = is_admin(request.user, school)
    staff = get_teacher_staff(request.user)

    hod_departments = get_hod_departments(staff)
    is_owner = bool(staff and section.teacher_id == staff.pk)
    hod_can_view = section_in_departments(section, hod_departments)

    if not (admin or is_owner or hod_can_view):
        messages.error(request, "Access denied.")
        return redirect("grades:home")

    term_filter = request.GET.get("term", str(section.term_number))
    try:
        term_filter = int(term_filter)
    except ValueError:
        term_filter = section.term_number

    all_sections = Section.objects.filter(
        school=school,
        course=section.course,
        form=section.form,
        academic_year=section.academic_year,
    ).order_by("term_number")

    active_section = all_sections.filter(term_number=term_filter).first() or section

    window_open = grade_window_is_open(school, active_section.academic_year, active_section.term_number, active_section.form)
    can_edit = admin or (is_owner and window_open)

    evaluations = Evaluation.objects.filter(school=school, section=active_section).order_by("date", "created_at")

    enrolments = active_section.enrolments.select_related("student").order_by("student__last_name", "student__first_name")
    students = [e.student for e in enrolments]

    entries = GradeEntry.objects.filter(
        school=school,
        evaluation__section=active_section,
    ).select_related("evaluation", "student")

    grade_map = {(e.evaluation_id, e.student_id): e for e in entries}

    comments = GradeComment.objects.filter(school=school, section=active_section)
    comment_map = {c.student_id: c.comment for c in comments}

    rows = []
    for student in students:
        cells = []
        for ev in evaluations:
            entry = grade_map.get((ev.pk, student.pk))
            cells.append(
                {
                    "ev": ev,
                    "entry": entry,
                    "pct": entry.percentage if entry else None,
                    "absent": entry.is_absent if entry else False,
                }
            )
        avg = compute_student_average(student, evaluations, grade_map)
        rows.append(
            {
                "student": student,
                "cells": cells,
                "avg": avg,
                "comment": comment_map.get(student.pk, ""),
            }
        )

    from scheduling.models import TermConfig

    term_configs = TermConfig.objects.filter(academic_year=active_section.academic_year)

    return render(
        request,
        "grades/section_table.html",
        {
            "section": active_section,
            "all_sections": all_sections,
            "term_filter": term_filter,
            "evaluations": evaluations,
            "rows": rows,
            "students": students,
            "can_edit": can_edit,
            "window_open": window_open,
            "is_admin": admin,
            "term_configs": term_configs,
            "is_hod_view": hod_can_view and not admin and not is_owner,
        },
    )
    pass


# -- Create Evaluation ---------------------------------------------------------


@login_required
@tenant_required
def evaluation_create(request, section_pk):
    section = get_object_or_404(Section, pk=section_pk, school=request.school)
    school = request.school
    admin = is_admin(request.user, school)
    staff = get_teacher_staff(request.user)

    if not admin and (not staff or section.teacher != staff):
        messages.error(request, "Access denied.")
        return redirect("grades:home")

    window_open = grade_window_is_open(school, section.academic_year, section.term_number, section.form)
    if not admin and not window_open:
        messages.error(request, "Grade window is closed. Contact admin to open it.")
        return redirect("grades:section_table", section_pk=section_pk)

    form = EvaluationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        ev = form.save(commit=False)
        ev.school = school
        ev.section = section
        ev.created_by = request.user
        ev.save()
        log_activity(request, "evaluation_created", f"Created evaluation '{ev.title}' in {section}.")
        messages.success(request, f"Evaluation '{ev.title}' created.")
        return redirect("grades:section_table", section_pk=section_pk)

    return render(
        request,
        "grades/evaluation_form.html",
        {
            "form": form,
            "section": section,
            "title": "Create Evaluation",
        },
    )
    pass


# -- Edit Evaluation -----------------------------------------------------------


@login_required
@tenant_required
def evaluation_edit(request, pk):
    ev = get_object_or_404(Evaluation, pk=pk, school=request.school)
    school = request.school
    admin = is_admin(request.user, school)
    staff = get_teacher_staff(request.user)

    if not admin and (not staff or ev.section.teacher != staff):
        messages.error(request, "Access denied.")
        return redirect("grades:home")

    form = EvaluationForm(request.POST or None, instance=ev)
    if request.method == "POST" and form.is_valid():
        form.save()
        log_activity(request, "evaluation_edited", f"Edited evaluation '{ev.title}' in {ev.section}.")
        messages.success(request, f"Evaluation '{ev.title}' updated.")
        return redirect("grades:section_table", section_pk=ev.section.pk)

    return render(
        request,
        "grades/evaluation_form.html",
        {
            "form": form,
            "section": ev.section,
            "title": f"Edit — {ev.title}",
            "ev": ev,
        },
    )
    pass


# -- Delete Evaluation ---------------------------------------------------------


@login_required
@tenant_required
def evaluation_delete(request, pk):
    ev = get_object_or_404(Evaluation, pk=pk, school=request.school)
    admin = is_admin(request.user, request.school)
    staff = get_teacher_staff(request.user)

    if not admin and (not staff or ev.section.teacher != staff):
        messages.error(request, "Access denied.")
        return redirect("grades:home")

    section_pk = ev.section.pk
    title = ev.title

    if request.method == "POST":
        ev.delete()
        log_activity(request, "evaluation_deleted", f"Deleted evaluation '{title}' from section pk={section_pk}.")
        messages.warning(request, f"Evaluation '{title}' and all its grades deleted.")
        return redirect("grades:section_table", section_pk=section_pk)

    return render(request, "grades/evaluation_confirm_delete.html", {"ev": ev})
    pass


# -- Save Grades ---------------------------------------------------------------


@login_required
@tenant_required
def grades_save(request, section_pk):
    section = get_object_or_404(Section, pk=section_pk, school=request.school)
    school = request.school
    admin = is_admin(request.user, school)
    staff = get_teacher_staff(request.user)

    if not admin and (not staff or section.teacher != staff):
        messages.error(request, "Access denied.")
        return redirect("grades:home")

    window_open = grade_window_is_open(school, section.academic_year, section.term_number, section.form)
    if not admin and not window_open:
        messages.error(request, "Grade window is closed.")
        return redirect("grades:section_table", section_pk=section_pk)

    if request.method != "POST":
        return redirect("grades:section_table", section_pk=section_pk)

    evaluations = Evaluation.objects.filter(school=school, section=section)
    enrolments = section.enrolments.select_related("student")
    students = [e.student for e in enrolments]
    saved = 0

    for student in students:
        for ev in evaluations:
            field_key = f"grade_{ev.pk}_{student.pk}"
            absent_key = f"absent_{ev.pk}_{student.pk}"
            raw = request.POST.get(field_key, "").strip()
            is_absent = request.POST.get(absent_key) == "on"

            marks = None
            if not is_absent and raw not in ("", "*", "null", "-"):
                try:
                    marks = Decimal(raw)
                    if marks < 0:
                        marks = Decimal("0")
                    if marks > ev.max_marks:
                        marks = ev.max_marks
                except InvalidOperation:
                    continue

            GradeEntry.objects.update_or_create(
                evaluation=ev,
                student=student,
                defaults={
                    "school": school,
                    "marks_earned": marks,
                    "is_absent": is_absent,
                    "entered_by": request.user,
                },
            )
            saved += 1

        comment_key = f"comment_{student.pk}"
        comment_text = request.POST.get(comment_key, "").strip()
        GradeComment.objects.update_or_create(
            section=section,
            student=student,
            defaults={
                "school": school,
                "comment": comment_text,
                "entered_by": request.user,
            },
        )

    log_activity(request, "grades_saved", f"Saved grades for {section} ({saved} entries).")
    messages.success(request, "Grades saved.")
    return redirect("grades:section_table", section_pk=section_pk)
    pass


# -- Bulk Grade Upload ---------------------------------------------------------


@login_required
@tenant_required
def bulk_grade_upload(request, section_pk):
    section = get_object_or_404(Section, pk=section_pk, school=request.school)
    school = request.school
    admin = is_admin(request.user, school)
    staff = get_teacher_staff(request.user)

    if not admin and (not staff or section.teacher != staff):
        messages.error(request, "Access denied.")
        return redirect("grades:home")

    window_open = grade_window_is_open(school, section.academic_year, section.term_number, section.form)
    if not admin and not window_open:
        messages.error(request, "Grade window is closed.")
        return redirect("grades:section_table", section_pk=section_pk)

    evaluations = Evaluation.objects.filter(school=school, section=section)
    form = BulkGradeUploadForm(request.POST or None, request.FILES or None)
    results = []
    errors = []

    if request.method == "POST" and "download_template" in request.POST:
        return _generate_grade_template(section, evaluations)

    if request.method == "POST" and form.is_valid():
        import openpyxl

        try:
            wb = openpyxl.load_workbook(request.FILES["excel_file"])
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Could not read file: {e}")
            return render(
                request,
                "grades/bulk_upload.html",
                {
                    "form": form,
                    "section": section,
                    "evaluations": evaluations,
                    "results": [],
                    "errors": [],
                },
            )

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        ev_map = {}
        for i, h in enumerate(headers):
            if h and str(h).startswith("EV_"):
                try:
                    ev_pk = int(str(h).replace("EV_", ""))
                    ev_obj = evaluations.filter(pk=ev_pk).first()
                    if ev_obj:
                        ev_map[i] = ev_obj
                except ValueError:
                    pass

        row_data = []
        for row_num in range(3, ws.max_row + 1):
            student_id = ws.cell(row=row_num, column=1).value
            if not student_id:
                continue
            student = Student.objects.filter(school=school, student_id=str(student_id).strip()).first()
            if not student:
                errors.append(f"Row {row_num}: Student ID '{student_id}' not found.")
                continue

            grades_row = {}
            for col_idx, ev in ev_map.items():
                val = ws.cell(row=row_num, column=col_idx + 1).value
                if val is None or str(val).strip() in ("*", "", "null", "-", "ABS"):
                    grades_row[ev.pk] = ("absent", None)
                else:
                    try:
                        marks = Decimal(str(val).strip())
                        if marks < 0:
                            marks = Decimal("0")
                        if marks > ev.max_marks:
                            marks = ev.max_marks
                        grades_row[ev.pk] = ("present", marks)
                    except InvalidOperation:
                        errors.append(f"Row {row_num}, {ev.title}: invalid value '{val}'.")
                        grades_row[ev.pk] = ("skip", None)

            row_data.append((student, grades_row))

        if errors:
            messages.error(request, f"Upload rejected — {len(errors)} error(s). Fix and re-upload.")
        else:
            for student, grades_row in row_data:
                for ev_pk, (status, marks) in grades_row.items():
                    if status == "skip":
                        continue
                    ev = evaluations.filter(pk=ev_pk).first()
                    if not ev:
                        continue
                    GradeEntry.objects.update_or_create(
                        evaluation=ev,
                        student=student,
                        defaults={
                            "school": school,
                            "marks_earned": marks,
                            "is_absent": status == "absent",
                            "entered_by": request.user,
                        },
                    )
                results.append(student.get_full_name())
            if results:
                log_activity(request, "grades_saved", f"Bulk uploaded grades for {len(results)} students in {section}.")
            messages.success(request, f"{len(results)} students updated.")

    return render(
        request,
        "grades/bulk_upload.html",
        {
            "form": form,
            "section": section,
            "evaluations": evaluations,
            "results": results,
            "errors": errors,
        },
    )
    pass


def _generate_grade_template(section, evaluations):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1e40af")
    sub_fill = PatternFill("solid", start_color="374151")
    sub_font = Font(name="Arial", size=8, color="FFFFFF")
    normal_font = Font(name="Arial", size=10)

    ws.cell(row=1, column=1, value="student_id").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2, value="student_name").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28

    for i, ev in enumerate(evaluations, start=3):
        col = get_column_letter(i)
        c = ws.cell(row=1, column=i, value=f"EV_{ev.pk}")
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

        label = ws.cell(row=2, column=i, value=f"{ev.title} (/{ev.max_marks})")
        label.font = sub_font
        label.fill = sub_fill
        label.alignment = Alignment(horizontal="center", wrap_text=True)

        ws.column_dimensions[col].width = 18
        ws.row_dimensions[2].height = 28

    ws.cell(row=2, column=1, value="Do not edit").font = sub_font
    ws.cell(row=2, column=1).fill = sub_fill
    ws.cell(row=2, column=2, value="Do not edit").font = sub_font
    ws.cell(row=2, column=2).fill = sub_fill

    enrolments = section.enrolments.select_related("student").order_by("student__last_name", "student__first_name")
    alt_fill = PatternFill("solid", start_color="F9FAFB")

    for row_num, enrolment in enumerate(enrolments, start=3):
        student = enrolment.student
        fill = alt_fill if row_num % 2 == 0 else None

        c1 = ws.cell(row=row_num, column=1, value=student.student_id)
        c1.font = Font(name="Arial", size=9)
        if fill:
            c1.fill = fill

        c2 = ws.cell(row=row_num, column=2, value=student.get_full_name())
        c2.font = Font(name="Arial", size=10)
        if fill:
            c2.fill = fill

        for i, ev in enumerate(evaluations, start=3):
            c = ws.cell(row=row_num, column=i, value="")
            c.font = normal_font
            c.alignment = Alignment(horizontal="center")
            if fill:
                c.fill = fill

    ws.freeze_panes = "C3"

    inst = wb.create_sheet("Instructions")
    inst["A1"] = "Grade Upload Instructions"
    inst["A1"].font = Font(name="Arial", bold=True, size=12)
    notes = [
        "1. Fill in grades in the white cells only (columns C onwards).",
        "2. Leave a cell blank, type *, ABS, or null to mark as ABSENT.",
        "   Absent students are excluded from average calculations.",
        "3. Do NOT edit columns A (student_id) or B (student_name).",
        "4. Do NOT change column header row 1.",
        "5. Marks are automatically capped at the maximum shown in row 2.",
        "6. Save as .xlsx before uploading.",
    ]
    for i, note in enumerate(notes, start=3):
        c = inst.cell(row=i, column=1, value=note)
        c.font = Font(name="Arial", size=10)
    inst.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="grades_{section.course.code or section.course.name}_term{section.term_number}.xlsx"'
    return response
    pass


# -- Grade Window Management ---------------------------------------------------


@login_required
@tenant_required
def grade_window_manage(request):
    if not is_admin(request.user, request.school):
        messages.error(request, "Access denied.")
        return redirect("portals:dashboard")

    school = request.school
    years = AcademicYear.objects.filter(school=school).order_by("-name")
    forms = school.forms.all()

    year_pk = request.GET.get("year") or (years.first().pk if years.exists() else None)
    year_pk = int(year_pk) if year_pk else None
    year = AcademicYear.objects.filter(pk=year_pk, school=school).first()

    from scheduling.models import TermConfig

    term_configs = TermConfig.objects.filter(academic_year=year) if year else []

    windows = {}
    if year:
        for w in GradeWindow.objects.filter(school=school, academic_year=year):
            windows[(w.form_id, w.term_number)] = w

    if request.method == "POST" and year:
        for form_obj in forms:
            for tc in term_configs:
                key = f"window_{form_obj.pk}_{tc.term_number}"
                is_open = request.POST.get(key) == "on"
                obj, _ = GradeWindow.objects.get_or_create(
                    school=school, academic_year=year, term_number=tc.term_number, form=form_obj, defaults={"is_open": is_open, "updated_by": request.user}
                )
                if obj.is_open != is_open:
                    obj.is_open = is_open
                    obj.updated_by = request.user
                    obj.save()

        log_activity(request, "grade_window_updated", f"Updated grade windows for {year} ({school}).")
        messages.success(request, "Grade windows updated.")
        return redirect(f"{request.path}?year={year_pk}")

    return render(
        request,
        "grades/grade_window.html",
        {
            "years": years,
            "year": year,
            "forms": forms,
            "term_configs": term_configs,
            "windows": windows,
            "is_admin": True,
        },
    )
    pass


# -- Visibility ----------------------------------------------------------------


@login_required
@tenant_required
def visibility_overview(request):
    from .models import GradeVisibilityRule
    from scheduling.models import Form, Homeroom

    school = request.school
    forms = Form.objects.filter(school=school)
    homerooms = Homeroom.objects.filter(school=school)

    form_pk = request.GET.get("form")
    homeroom_pk = request.GET.get("homeroom")

    students = None
    selected_form = None
    selected_homeroom = None

    if form_pk:
        from students.models import Student

        selected_form = forms.filter(pk=form_pk).first()
        students = Student.objects.filter(school=school, form_id=form_pk).order_by("last_name", "first_name")

    elif homeroom_pk:
        from students.models import Student

        selected_homeroom = homerooms.filter(pk=homeroom_pk).first()
        students = Student.objects.filter(school=school, homeroom_id=homeroom_pk).order_by("last_name", "first_name")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "toggle_all_on":
            from students.models import Student as S

            all_students = S.objects.filter(school=school)
            for student in all_students:
                GradeVisibilityRule.objects.update_or_create(school=school, student=student, defaults={"is_visible": True, "set_by": request.user})
            GradeVisibilityRule.objects.update_or_create(school=school, student=None, defaults={"is_visible": True, "set_by": request.user})
            messages.success(request, "Grades visible for all students.")
            return redirect(request.get_full_path())

        if action == "toggle_all_off":
            from students.models import Student as S

            all_students = S.objects.filter(school=school)
            for student in all_students:
                GradeVisibilityRule.objects.update_or_create(school=school, student=student, defaults={"is_visible": False, "set_by": request.user})
            GradeVisibilityRule.objects.update_or_create(school=school, student=None, defaults={"is_visible": False, "set_by": request.user})
            messages.success(request, "Grades hidden for all students.")
            return redirect(request.get_full_path())

        if action == "save_students":
            if students:
                for student in students:
                    val = request.POST.get(f"visible_{student.pk}") == "on"
                    GradeVisibilityRule.objects.update_or_create(school=school, student=student, defaults={"is_visible": val, "set_by": request.user})
                messages.success(request, "Visibility updated.")
            return redirect(request.get_full_path())

    visibility_map = {}
    if students:
        rules = GradeVisibilityRule.objects.filter(school=school, student__in=students)
        visibility_map = {r.student_id: r.is_visible for r in rules}

    school_rule = GradeVisibilityRule.objects.filter(school=school, student__isnull=True).order_by("-updated_at").first()

    return render(
        request,
        "grades/visibility_overview.html",
        {
            "forms": forms,
            "homerooms": homerooms,
            "students": students,
            "visibility_map": visibility_map,
            "selected_form": selected_form,
            "selected_homeroom": selected_homeroom,
            "school_rule": school_rule,
            "form_pk": form_pk,
            "homeroom_pk": homeroom_pk,
        },
    )
    pass


@login_required
@tenant_required
def visibility_set_school(request):
    from .models import GradeVisibilityRule

    existing = GradeVisibilityRule.objects.filter(school=request.school, student__isnull=True).order_by("-updated_at").first()
    form = GradeVisibilityForm(request.POST or None, instance=existing)
    if request.method == "POST" and form.is_valid():
        rule = form.save(commit=False)
        rule.school = request.school
        rule.student = None
        rule.set_by = request.user
        if existing:
            existing.is_visible = rule.is_visible
            existing.reason = rule.reason
            existing.set_by = request.user
            existing.save()
        else:
            rule.save()
        messages.success(request, "School-wide visibility updated.")
        return redirect("grades:visibility")
    return render(
        request,
        "grades/visibility_form.html",
        {
            "form": form,
            "title": "School-wide Grade Visibility",
        },
    )
    pass


@login_required
@tenant_required
def visibility_set_student(request, student_pk):
    from .models import GradeVisibilityRule

    student = get_object_or_404(Student, pk=student_pk, school=request.school)
    existing = GradeVisibilityRule.objects.filter(school=request.school, student=student).order_by("-updated_at").first()
    form = GradeVisibilityForm(request.POST or None, instance=existing)
    if request.method == "POST" and form.is_valid():
        rule = form.save(commit=False)
        rule.school = request.school
        rule.student = student
        rule.set_by = request.user
        if existing:
            existing.is_visible = rule.is_visible
            existing.reason = rule.reason
            existing.set_by = request.user
            existing.save()
        else:
            rule.save()
        messages.success(request, f"Visibility updated for {student.get_full_name()}.")
        return redirect("grades:visibility")
    return render(
        request,
        "grades/visibility_form.html",
        {
            "form": form,
            "title": f"Grade Visibility — {student.get_full_name()}",
            "student": student,
        },
    )
    pass


# -- Report Cards --------------------------------------------------------------


@login_required
@tenant_required
def report_card_list(request):
    year_pk = request.GET.get("year")
    term = request.GET.get("term")
    report_cards = ReportCard.objects.filter(school=request.school).select_related("student", "academic_year")
    years = AcademicYear.objects.filter(school=request.school)
    if year_pk:
        report_cards = report_cards.filter(academic_year_id=year_pk)
    if term:
        report_cards = report_cards.filter(term_number=term)
    return render(
        request,
        "grades/report_card_list.html",
        {
            "report_cards": report_cards,
            "years": years,
            "selected_year": year_pk,
            "selected_term": term,
        },
    )
    pass


@login_required
@tenant_required
def report_card_detail(request, pk):
    rc = get_object_or_404(ReportCard, pk=pk, school=request.school)
    enrolments = Enrolment.objects.filter(
        student=rc.student,
        section__academic_year=rc.academic_year,
        section__term_number=rc.term_number,
        section__school=request.school,
    ).select_related("section__course")

    rows = []
    for enrolment in enrolments:
        section = enrolment.section
        evs = Evaluation.objects.filter(school=request.school, section=section)
        ents = GradeEntry.objects.filter(evaluation__section=section, student=rc.student)
        gmap = {(e.evaluation_id, e.student_id): e for e in ents}

        course_evs = [e for e in evs if not e.is_final_exam]
        exam_evs = [e for e in evs if e.is_final_exam]

        coursework_avg = compute_student_average(rc.student, course_evs, gmap)
        exam_mark = compute_student_average(rc.student, exam_evs, gmap) if exam_evs else None
        has_exam = bool(exam_evs)

        parts = [v for v in (coursework_avg, exam_mark) if v is not None]
        term_grade = round(sum(parts) / len(parts), 1) if parts else None

        result = None
        if coursework_avg is not None or exam_mark is not None:
            result = {
                "coursework_avg": coursework_avg if coursework_avg is not None else "—",
                "exam_mark": exam_mark,
                "has_exam": has_exam,
                "term_grade": term_grade if term_grade is not None else "—",
            }

        rows.append({"course": section.course, "result": result})

    return render(request, "grades/report_card_detail.html", {"rc": rc, "rows": rows})
    pass
