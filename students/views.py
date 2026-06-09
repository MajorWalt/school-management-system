import csv
import io
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from core.decorators import tenant_required
from core.activity import log_activity
from accounts.models import UserRole
from scheduling.models import Form, Homeroom
from .forms import BulkEnrolForm, GuardianForm, StudentForm, StudentGuardianForm, StudentStatusForm
from .models import Guardian, House, Student, StudentGuardian, StudentStatusLog


def is_admin(user, school):
	return UserRole.objects.filter(user=user, school=school, role="admin").exists() or user.is_superuser


@login_required
@tenant_required
def student_list(request):
	form_pk     = request.GET.get("form")
	homeroom_pk = request.GET.get("homeroom")
	show_all    = request.GET.get("all")
	query       = request.GET.get("q", "")
	status      = request.GET.get("status", "enrolled")

	forms      = Form.objects.filter(school=request.school)
	homerooms  = Homeroom.objects.filter(school=request.school)
	students   = None
	selected_form     = None
	selected_homeroom = None

	if form_pk or homeroom_pk or show_all:
		students = Student.objects.filter(school=request.school)

		if form_pk:
			students      = students.filter(form_id=form_pk)
			selected_form = forms.filter(pk=form_pk).first()

		if homeroom_pk:
			students          = students.filter(homeroom_id=homeroom_pk)
			selected_homeroom = homerooms.filter(pk=homeroom_pk).first()

		if query:
			terms    = query.split()
			q_filter = Q()
			for term in terms:
				q_filter |= (
					Q(first_name__icontains=term)  |
					Q(last_name__icontains=term)   |
					Q(middle_name__icontains=term) |
					Q(student_id__icontains=term)
				)
			students = students.filter(q_filter).distinct()

		if status and status != "all":
			enrolled_ids = [s.pk for s in students if s.current_status() == status]
			students     = students.filter(pk__in=enrolled_ids)

	return render(request, "students/student_list.html", {
		"students":          students,
		"query":             query,
		"status":            status,
		"forms":             forms,
		"homerooms":         homerooms,
		"selected_form":     selected_form,
		"selected_homeroom": selected_homeroom,
		"show_all":          show_all,
		"is_admin":          is_admin(request.user, request.school),
	})


@login_required
@tenant_required
def student_detail(request, pk):
	student   = get_object_or_404(Student, pk=pk, school=request.school)
	logs      = student.status_logs.order_by("-change_date")
	guardians = student.guardians.select_related("guardian")
	return render(request, "students/student_detail.html", {
		"student":  student,
		"logs":     logs,
		"guardians": guardians,
		"is_admin": is_admin(request.user, request.school),
	})


@login_required
@tenant_required
def student_add(request):
	if not is_admin(request.user, request.school):
		messages.error(request, "Access denied.")
		return redirect("students:list")

	form = StudentForm(request.POST or None, request.FILES or None, school=request.school)
	if request.method == "POST" and form.is_valid():
		student = form.save(commit=False)
		student.school = request.school
		student.save()
		StudentStatusLog.objects.create(
			student     = student,
			status      = "enrolled",
			change_date = student.admission_date or student.created_at.date(),
			changed_by  = request.user,
			reason      = "Initial enrolment",
		)
		log_activity(request, "student_add", f"Enrolled student {student.get_full_name()} ({student.student_id}).")
		messages.success(request, f"{student.get_full_name()} added successfully.")
		return redirect("students:detail", pk=student.pk)
	return render(request, "students/student_form.html", {"form": form, "title": "Enrol Student"})

@login_required
@tenant_required
def student_bulk_enrol(request):
	if not is_admin(request.user, request.school):
		messages.error(request, "Access denied.")
		return redirect("students:list")

	form    = BulkEnrolForm(request.POST or None, request.FILES or None)
	results = []
	errors  = []

	if request.method == "POST" and form.is_valid():
		uploaded = request.FILES["excel_file"]

		try:
			import openpyxl
			wb     = openpyxl.load_workbook(uploaded)
			ws     = wb["Students"]
		except Exception as e:
			messages.error(request, f"Could not read file: {e}")
			return render(request, "students/bulk_enrol.html", {"form": form, "results": [], "errors": []})

		# Read headers from row 1
		headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

		# Validate all rows first — collect all errors before touching the DB
		rows_data = []
		for row_num in range(3, ws.max_row + 1):
			row = {
				headers[c]: ws.cell(row=row_num, column=c + 1).value
				for c in range(len(headers))
				if headers[c]
			}

			# Skip completely empty rows
			if not any(v for v in row.values() if v is not None and str(v).strip()):
				continue

			row_errors = []
			sid        = str(row.get("student_id", "") or "").strip()
			first_name = str(row.get("first_name", "") or "").strip()
			last_name  = str(row.get("last_name", "") or "").strip()
			gender     = str(row.get("gender", "") or "").strip().upper()

			if not sid:
				row_errors.append("student_id is required")
			if not first_name:
				row_errors.append("first_name is required")
			if not last_name:
				row_errors.append("last_name is required")
			if not gender:
				row_errors.append("gender is required")
			elif gender not in ("M", "F"):
				row_errors.append(f"gender must be M or F, got '{gender}'")

			# Check duplicate within file
			if sid and any(r["student_id"] == sid for r in rows_data):
				row_errors.append(f"Duplicate student_id '{sid}' in this file")

			# Check duplicate in DB
			if sid and Student.objects.filter(school=request.school, student_id=sid).exists():
				row_errors.append(f"student_id '{sid}' already exists in the system")

			# Validate date
			dob_str = str(row.get("date_of_birth", "") or "").strip()
			dob     = None
			if dob_str:
				from datetime import date
				try:
					dob = date.fromisoformat(dob_str)
				except ValueError:
					row_errors.append(f"date_of_birth '{dob_str}' must be YYYY-MM-DD format")

			# Validate homeroom
			hr_name  = str(row.get("homeroom", "") or "").strip()
			homeroom = None
			if hr_name:
				homeroom = Homeroom.objects.filter(school=request.school, name=hr_name).first()
				if not homeroom:
					row_errors.append(f"Homeroom '{hr_name}' not found in system")

			if row_errors:
				errors.append({
					"row":    row_num - 2,
					"sid":    sid or "—",
					"name":   f"{first_name} {last_name}".strip() or "—",
					"errors": row_errors,
				})
			else:
				rows_data.append({
					"student_id":             sid,
					"first_name":             first_name,
					"last_name":              last_name,
					"middle_name":            str(row.get("middle_name", "") or "").strip(),
					"gender":                 gender,
					"date_of_birth":          dob,
					"homeroom":               homeroom,
					"form":                   homeroom.form if homeroom else None,
					"nationality":            str(row.get("nationality", "") or "").strip(),
					"religion":               str(row.get("religion", "") or "").strip(),
					"phone":                  str(row.get("phone", "") or "").strip(),
					"email":                  str(row.get("email", "") or "").strip(),
					"address":                str(row.get("address", "") or "").strip(),
					"city":                   str(row.get("city", "") or "").strip(),
					"parish":                 str(row.get("parish", "") or "").strip(),
					"community":              str(row.get("community", "") or "").strip(),
					"father_name":            str(row.get("father_name", "") or "").strip(),
					"mother_name":            str(row.get("mother_name", "") or "").strip(),
					"emergency_contact_name": str(row.get("emergency_contact_name", "") or "").strip(),
					"emergency_relation":     str(row.get("emergency_relation", "") or "").strip(),
					"emergency_phone_1":      str(row.get("emergency_phone_1", "") or "").strip(),
					"emis_id":                str(row.get("emis_id", "") or "").strip(),
					"previous_school":        str(row.get("previous_school", "") or "").strip(),
					"notes":                  str(row.get("notes", "") or "").strip(),
				})

		# If ANY errors — reject entire upload
		if errors:
			messages.error(request, f"Upload rejected — {len(errors)} error(s) found. Fix them and re-upload.")
			return render(request, "students/bulk_enrol.html", {
				"form":    form,
				"results": [],
				"errors":  errors,
			})

		# All rows valid — now create
		from datetime import date as date_cls
		created_count = 0
		for data in rows_data:
			student = Student.objects.create(
				school         = request.school,
				admission_date = date_cls.today(),
				**data,
			)
			StudentStatusLog.objects.create(
				student     = student,
				status      = "enrolled",
				change_date = date_cls.today(),
				changed_by  = request.user,
				reason      = "Bulk enrolment",
			)
			results.append({"name": student.get_full_name(), "sid": student.student_id})
			created_count += 1

		log_activity(request, "student_bulk_enrol", f"Bulk enrolled {created_count} students.")
		messages.success(request, f"{created_count} students enrolled successfully.")

	return render(request, "students/bulk_enrol.html", {
		"form":    form,
		"results": results,
		"errors":  errors,
	})


@login_required
@tenant_required
def student_edit(request, pk):
	if not is_admin(request.user, request.school):
		messages.error(request, "Access denied.")
		return redirect("students:list")

	student = get_object_or_404(Student, pk=pk, school=request.school)
	form    = StudentForm(request.POST or None, request.FILES or None, instance=student, school=request.school)
	if request.method == "POST" and form.is_valid():
		form.save()
		log_activity(request, "student_edit", f"Edited student {student.get_full_name()} ({student.student_id}).")
		messages.success(request, f"{student.get_full_name()} updated successfully.")
		return redirect("students:detail", pk=pk)
	return render(request, "students/student_form.html", {"form": form, "title": "Edit Student"})


@login_required
@tenant_required
def student_status_change(request, pk):
	student = get_object_or_404(Student, pk=pk, school=request.school)
	form    = StudentStatusForm(request.POST or None)
	if request.method == "POST" and form.is_valid():
		log            = form.save(commit=False)
		log.student    = student
		log.changed_by = request.user
		log.save()
		log_activity(request, "student_status", f"Changed status of {student.get_full_name()} to {log.get_status_display()}.")
		messages.success(request, f"Status updated to {log.get_status_display()}.")
		return redirect("students:detail", pk=pk)
	return render(request, "students/student_status_form.html", {
		"form":    form,
		"student": student,
	})


@login_required
@tenant_required
def guardian_add(request, student_pk):
	student = get_object_or_404(Student, pk=student_pk, school=request.school)
	g_form  = GuardianForm(request.POST or None)
	sg_form = StudentGuardianForm(request.POST or None, school=request.school)

	if request.method == "POST" and g_form.is_valid() and sg_form.is_valid():
		guardian        = g_form.save(commit=False)
		guardian.school = request.school
		guardian.save()
		link          = sg_form.save(commit=False)
		link.student  = student
		link.guardian = guardian
		link.save()
		messages.success(request, f"Guardian {guardian.get_full_name()} added.")
		return redirect("students:detail", pk=student_pk)

	return render(request, "students/guardian_form.html", {
		"g_form":  g_form,
		"sg_form": sg_form,
		"student": student,
	})

from .forms import BulkEnrolForm, GuardianForm, StudentForm, StudentGuardianForm, StudentStatusForm, WithdrawForm


@login_required
@tenant_required
def student_withdraw(request, pk):
	if not is_admin(request.user, request.school):
		messages.error(request, "Access denied.")
		return redirect("students:list")

	student = get_object_or_404(Student, pk=pk, school=request.school)

	# Already withdrawn
	if student.current_status() == "withdrawn":
		messages.warning(request, f"{student.get_full_name()} is already withdrawn.")
		return redirect("students:detail", pk=pk)

	import datetime
	form = WithdrawForm(
		request.POST or None,
		initial={"withdraw_date": datetime.date.today().isoformat()}
	)

	if request.method == "POST" and form.is_valid():
		StudentStatusLog.objects.create(
			student     = student,
			status      = "withdrawn",
			change_date = form.cleaned_data["withdraw_date"],
			reason      = form.cleaned_data["reason"],
			description = form.cleaned_data["description"],
			changed_by  = request.user,
		)
		messages.success(request, f"{student.get_full_name()} has been withdrawn.")
		return redirect("students:detail", pk=pk)

	return render(request, "students/withdraw.html", {
		"form":    form,
		"student": student,
	})